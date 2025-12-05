import re
import pandas as pd
from collections import defaultdict
from flask import Flask, request, render_template, send_file, redirect, url_for
from io import BytesIO
from datetime import datetime
import pytz
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import sys
import gc
import json
import time
import claim_processor  # Import the new module

app = Flask(__name__)

# ---------------------------------------------------------
# SHARED / DATA MAPPING LOGIC
# ---------------------------------------------------------

sku_category_mapping = {
    "Warranty : Water Cooler/Dispencer/Geyser/RoomCooler/Heater": [
        "COOLER", "DISPENCER", "GEYSER", "ROOM COOLER", "HEATER", "WATER HEATER", "WATER DISPENSER"
    ],
    "Warranty : Fan/Mixr/IrnBox/Kettle/OTG/Grmr/Geysr/Steamr/Inductn": [
        "FAN", "MIXER", "IRON BOX", "KETTLE", "OTG", "GROOMING KIT", "GEYSER", "STEAMER", "INDUCTION",
        "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "INDUCTION COOKER", "ELECTRIC KETTLE", "WALL FAN", "MIXER GRINDER", "CELLING FAN"
    ],
    "AC : EWP : Warranty : AC": ["AC", "AIR CONDITIONER", "AC INDOOR"],
    "HAEW : Warranty : Air Purifier/WaterPurifier": ["AIR PURIFIER", "WATER PURIFIER"],
    "HAEW : Warranty : Dryer/MW/DishW": ["DRYER", "MICROWAVE OVEN", "DISH WASHER", "MICROWAVE OVEN-CONV"],
    "HAEW : Warranty : Ref/WM": [
        "REFRIGERATOR", "WASHING MACHINE", "WASHING MACHINE-TL", "REFRIGERATOR-DC",
        "WASHING MACHINE-FL", "WASHING MACHINE-SA", "REF", "REFRIGERATOR-CBU", "REFRIGERATOR-FF", "WM"
    ],
    "HAEW : Warranty : TV": ["TV", "TV 28 %", "TV 18 %"],
    "TV : TTC : Warranty and Protection : TV": ["TV", "TV 28 %", "TV 18 %"],
    "TV : Spill and Drop Protection": ["TV", "TV 28 %", "TV 18 %"],
    "HAEW : Warranty :Chop/Blend/Toast/Air Fryer/Food Processr/JMG/Induction": [
        "CHOPPER", "BLENDER", "TOASTER", "AIR FRYER", "FOOD PROCESSOR", "JUICER", "INDUCTION COOKER"
    ],
    "HAEW : Warranty : HOB and Chimney": ["HOB", "CHIMNEY"],
    "HAEW : Warranty : HT/SoundBar/AudioSystems/PortableSpkr": [
        "HOME THEATRE", "AUDIO SYSTEM", "SPEAKER", "SOUND BAR", "PARTY SPEAKER"
    ],
    "HAEW : Warranty : Vacuum Cleaner/Fans/Groom&HairCare/Massager/Iron": [
        "VACUUM CLEANER", "FAN", "MASSAGER", "IRON BOX", "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "WALL FAN", "ROBO VACCUM CLEANER"
    ],
    "AC AMC": ["AC", "AC INDOOR"]
}

def extract_price_slab(text):
    match = re.search(r"Slab\s*:\s*(\d+)K-(\d+)K", str(text))
    if match:
        return int(match.group(1)) * 1000, int(match.group(2)) * 1000
    return None, None

def extract_warranty_duration(sku):
    sku = str(sku)
    match = re.search(r'Dur\s*:\s*(\d+)\+(\d+)', sku)
    if match:
        return int(match.group(1)), int(match.group(2))
    match = re.search(r'(\d+)\+(\d+)\s*SDP-(\d+)', sku)
    if match:
        return int(match.group(1)), f"{match.group(3)}P+{match.group(2)}W"
    match = re.search(r'Dur\s*:\s*(\d+)', sku)
    if match:
        return 1, int(match.group(1))
    match = re.search(r'(\d+)\+(\d+)', sku)
    if match:
        return int(match.group(1)), int(match.group(2))
    return '', ''

def highlight_row(row):
    missing_fields = pd.isna(row.get('Model')) or str(row.get('Model')).strip() == ''
    missing_fields |= pd.isna(row.get('IMEI')) or str(row.get('IMEI')).strip() == ''
    try:
        if float(row.get('Plan Price', 0)) < 0:
            missing_fields |= True
    except:
        missing_fields |= True
    return ['background-color: lightblue'] * len(row) if missing_fields else [''] * len(row)

# ---------------------------------------------------------
# ROUTES
# ---------------------------------------------------------

@app.route("/")
def index():
    return redirect(url_for('mapping_page'))

@app.route("/health")
def health():
    return {"status": "ok", "version": "4.0", "deployed": "2025-12-02 15:10"}

@app.route("/mapping")
def mapping_page():
    return render_template("mapping.html")

@app.route("/report1")
def report1_page():
    return render_template("report1.html")

@app.route("/report2")
def report2_page():
    return render_template("report2.html")

# ---------------------------------------------------------
# PROCESS: REPORT 1 (SALES REPORT) - STREAMLIT LOGIC PORT
# ---------------------------------------------------------

@app.route("/process_report1", methods=["POST"])
def process_report1():
    try:
        print("=== START REPORT 1 PROCESSING (v4.0 - Streamlit Logic Port) ===", file=sys.stderr)
        
        # Check for xlsxwriter
        try:
            import xlsxwriter
        except ImportError:
            return "ERROR: xlsxwriter module is missing on the server.", 500

        report_date = pd.to_datetime(request.form['report_date'])
        prev_date = pd.to_datetime(request.form['prev_date'])
        
        curr_osg_file = request.files['curr_osg_file']
        product_file = request.files['product_file']
        prev_osg_file = request.files.get('prev_osg_file')
        
        # Load Master Files
        try:
            future_store_df = pd.read_excel("myG All Store.xlsx", engine='openpyxl')
            rbm_df = pd.read_excel("RBM,BDM,BRANCH.xlsx", engine='openpyxl')
            print("Loaded master files.", file=sys.stderr)
        except Exception as e:
            return f"Error loading master files: {e}", 500

        # Process OSG File
        book1_df = pd.read_excel(curr_osg_file, engine='openpyxl')
        
        # Robust Column Normalization for OSG
        osg_renames = {}
        for col in book1_df.columns:
            c = str(col).strip().lower()
            if c in ['date']: osg_renames[col] = 'DATE'
            elif c in ['quantity', 'qty'] and 'billed' not in c: osg_renames[col] = 'QUANTITY'
            elif 'billed' in c: osg_renames[col] = 'BILLED_QTY'
            elif c in ['amount']: osg_renames[col] = 'AMOUNT'
            elif c in ['branch', 'store']: osg_renames[col] = 'Store'
        
        # Priority check for Store
        store_found = False
        for p in ['store', 'store name', 'branch', 'branch name', 'outlet']:
            for col in book1_df.columns:
                if str(col).strip().lower() == p:
                    osg_renames[col] = 'Store'
                    store_found = True
                    break
            if store_found: break
            
        if osg_renames: book1_df.rename(columns=osg_renames, inplace=True)
        
        # Handle Quantity/Billed Qty
        if 'QUANTITY' not in book1_df.columns:
            if 'BILLED_QTY' in book1_df.columns: book1_df['QUANTITY'] = book1_df['BILLED_QTY']
            else: book1_df['QUANTITY'] = 1

        book1_df['DATE'] = pd.to_datetime(book1_df['DATE'], dayfirst=True, errors='coerce')
        book1_df = book1_df.dropna(subset=['DATE'])
        
        # Normalize RBM DF
        rbm_renames = {}
        for col in rbm_df.columns:
            c = str(col).strip().lower()
            if c in ['branch', 'store']: rbm_renames[col] = 'Store'
            elif c in ['rbm', 'manager']: rbm_renames[col] = 'RBM'
        if rbm_renames: rbm_df.rename(columns=rbm_renames, inplace=True)

        # Process Product File
        product_df = pd.read_excel(product_file, engine='openpyxl')
        prod_renames = {}
        for col in product_df.columns:
            c = str(col).strip().lower()
            if c in ['branch', 'store']: prod_renames[col] = 'Store'
            elif c in ['date']: prod_renames[col] = 'DATE'
            elif c in ['sold price', 'amount', 'price']: prod_renames[col] = 'AMOUNT'
            elif c in ['quantity', 'qty']: prod_renames[col] = 'QUANTITY'
        if prod_renames: product_df.rename(columns=prod_renames, inplace=True)
        
        product_df['DATE'] = pd.to_datetime(product_df['DATE'], dayfirst=True, errors='coerce')
        product_df = product_df.dropna(subset=['DATE'])
        if 'QUANTITY' not in product_df.columns: product_df['QUANTITY'] = 1

        # Aggregations
        today = pd.to_datetime(report_date)
        mtd_df = book1_df[book1_df['DATE'].dt.month == today.month]
        today_df = mtd_df[mtd_df['DATE'].dt.date == today.date()]
        
        today_agg = today_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'FTD Count', 'AMOUNT': 'FTD Value'})
        mtd_agg = mtd_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'MTD Count', 'AMOUNT': 'MTD Value'})

        product_mtd_df = product_df[product_df['DATE'].dt.month == today.month]
        product_today_df = product_mtd_df[product_mtd_df['DATE'].dt.date == today.date()]
        
        product_today_agg = product_today_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'Product_FTD_Count', 'AMOUNT': 'Product_FTD_Amount'})
        product_mtd_agg = product_mtd_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'Product_MTD_Count', 'AMOUNT': 'Product_MTD_Amount'})

        # Previous Month
        if prev_osg_file and prev_osg_file.filename != '':
            prev_df = pd.read_excel(prev_osg_file, engine='openpyxl')
            p_renames = {}
            for col in prev_df.columns:
                c = str(col).strip().lower()
                if c in ['branch', 'store']: p_renames[col] = 'Store'
                elif c in ['date']: p_renames[col] = 'DATE'
                elif c in ['amount']: p_renames[col] = 'AMOUNT'
            if p_renames: prev_df.rename(columns=p_renames, inplace=True)
            
            prev_df['DATE'] = pd.to_datetime(prev_df['DATE'], dayfirst=True, errors='coerce')
            prev_df = prev_df.dropna(subset=['DATE'])
            prev_month = pd.to_datetime(prev_date)
            prev_mtd_df = prev_df[prev_df['DATE'].dt.month == prev_month.month]
            prev_mtd_agg = prev_mtd_df.groupby('Store', as_index=False).agg({'AMOUNT': 'sum'}).rename(columns={'AMOUNT': 'PREV MONTH SALE'})
        else:
            prev_mtd_agg = pd.DataFrame(columns=['Store', 'PREV MONTH SALE'])

        # Merge
        for col in future_store_df.columns:
            if str(col).strip().lower() in ['store', 'branch']:
                future_store_df.rename(columns={col: 'Store'}, inplace=True)
                break

        all_stores_list = pd.concat([future_store_df['Store'], book1_df['Store'], product_df['Store']]).unique()
        all_stores = pd.DataFrame(all_stores_list, columns=['Store'])
        
        report_df = all_stores.merge(today_agg, on='Store', how='left') \
                              .merge(mtd_agg, on='Store', how='left') \
                              .merge(product_today_agg, on='Store', how='left') \
                              .merge(product_mtd_agg, on='Store', how='left') \
                              .merge(prev_mtd_agg, on='Store', how='left') \
                              .merge(rbm_df.rename(columns={'Branch': 'Store'})[['Store', 'RBM']], on='Store', how='left')


        # Fill NaNs
        required_columns = ['Store', 'FTD Count', 'FTD Value', 'Product_FTD_Amount', 'MTD Count', 'MTD Value', 'Product_MTD_Amount', 'PREV MONTH SALE', 'RBM']
        for col in required_columns:
            if col not in report_df.columns: report_df[col] = 0
            
        report_df = report_df.rename(columns={'Store': 'Store Name'})
        
        cols_to_int = ['FTD Count', 'FTD Value', 'MTD Count', 'MTD Value', 'Product_FTD_Count', 'Product_FTD_Amount', 'Product_MTD_Count', 'Product_MTD_Amount', 'PREV MONTH SALE']
        report_df[cols_to_int] = report_df[cols_to_int].fillna(0).astype(int)

        # Metrics
        report_df['DIFF %'] = report_df.apply(lambda x: round(((x['MTD Value'] - x['PREV MONTH SALE']) / x['PREV MONTH SALE']) * 100, 2) if x['PREV MONTH SALE'] != 0 else 0, axis=1)
        report_df['ASP'] = report_df.apply(lambda x: round(x['MTD Value'] / x['MTD Count'], 2) if x['MTD Count'] != 0 else 0, axis=1)
        report_df['FTD Value Conversion'] = report_df.apply(lambda x: round((x['FTD Value'] / x['Product_FTD_Amount']) * 100, 2) if x['Product_FTD_Amount'] != 0 else 0, axis=1)
        report_df['MTD Value Conversion'] = report_df.apply(lambda x: round((x['MTD Value'] / x['Product_MTD_Amount']) * 100, 2) if x['Product_MTD_Amount'] != 0 else 0, axis=1)

        # Excel Generation
        excel_output = BytesIO()
        with pd.ExcelWriter(excel_output, engine='xlsxwriter') as writer:
            workbook = writer.book
            
            colors_palette = {
                'primary_blue': '#1E3A8A', 'light_blue': '#DBEAFE', 'success_green': '#065F46', 'light_green': '#D1FAE5',
                'warning_orange': '#EA580C', 'light_orange': '#FED7AA', 'danger_red': '#DC2626', 'light_red': '#FEE2E2',
                'accent_purple': '#7C3AED', 'light_purple': '#EDE9FE', 'neutral_gray': '#6B7280', 'light_gray': '#F9FAFB',
                'white': '#FFFFFF', 'dark_blue': '#0F172A', 'mint_green': '#10B981', 'light_mint': '#ECFDF5',
                'royal_blue': '#3B82F6', 'light_royal': '#EBF8FF'
            }

            formats = {
                'title': workbook.add_format({'bold': True, 'font_size': 16, 'font_color': colors_palette['primary_blue'], 'align': 'center', 'valign': 'vcenter', 'bg_color': colors_palette['white'], 'border': 1, 'border_color': colors_palette['primary_blue']}),
                'subtitle': workbook.add_format({'bold': True, 'font_size': 12, 'font_color': colors_palette['neutral_gray'], 'align': 'center', 'valign': 'vcenter', 'italic': True}),
                'header_main': workbook.add_format({'bold': True, 'font_size': 11, 'font_color': colors_palette['white'], 'bg_color': colors_palette['primary_blue'], 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['primary_blue'], 'text_wrap': True}),
                'header_secondary': workbook.add_format({'bold': True, 'font_size': 10, 'font_color': colors_palette['primary_blue'], 'bg_color': colors_palette['light_blue'], 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['primary_blue']}),
                'data_normal': workbook.add_format({'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white']}),
                'data_alternate': workbook.add_format({'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_gray']}),
                'data_store_name': workbook.add_format({'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white'], 'indent': 1}),
                'data_store_name_alt': workbook.add_format({'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_gray'], 'indent': 1}),
                'conversion_low': workbook.add_format({'font_size': 10, 'font_color': colors_palette['danger_red'], 'bg_color': colors_palette['light_red'], 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['danger_red'], 'num_format': '0.00%', 'bold': True}),
                'conversion_green': workbook.add_format({'bold': True, 'font_size': 10, 'font_color': colors_palette['success_green'], 'bg_color': colors_palette['light_green'], 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['success_green'], 'num_format': '0.00%'}),
                'conversion_format': workbook.add_format({'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'num_format': '0.00%'}),
                'conversion_format_alt': workbook.add_format({'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'num_format': '0.00%'}),
                'total_row': workbook.add_format({'bold': True, 'font_size': 11, 'font_color': colors_palette['white'], 'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter', 'border': 2, 'border_color': colors_palette['mint_green']}),
                'total_label': workbook.add_format({'bold': True, 'font_size': 11, 'font_color': colors_palette['white'], 'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter', 'border': 2, 'border_color': colors_palette['mint_green']}),
                'asp_format': workbook.add_format({'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1,  'border_color': colors_palette['neutral_gray'], 'num_format': '₹#,##0.00'}),
                'asp_format_alt': workbook.add_format({'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'num_format': '₹#,##0.00'}),
                'asp_total': workbook.add_format({'bold': True, 'font_size': 12, 'font_color': colors_palette['white'], 'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter', 'border': 2, 'border_color': colors_palette['mint_green'], 'num_format': '₹#,##0.00'}),
                'rbm_title': workbook.add_format({'bold': True, 'font_size': 18, 'font_color': colors_palette['white'], 'bg_color': colors_palette['dark_blue'], 'align': 'center', 'valign': 'vcenter', 'border': 2, 'border_color': colors_palette['dark_blue']}),
                'rbm_subtitle': workbook.add_format({'bold': True, 'font_size': 11, 'font_color': colors_palette['dark_blue'], 'bg_color': colors_palette['light_royal'], 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['royal_blue'], 'italic': True}),
                'rbm_header': workbook.add_format({'bold': True, 'font_size': 11, 'font_color': colors_palette['white'], 'bg_color': colors_palette['royal_blue'], 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['royal_blue'], 'text_wrap': True}),
                'rbm_data_normal': workbook.add_format({'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white']}),
                'rbm_data_alternate': workbook.add_format({'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal']}),
                'rbm_store_name': workbook.add_format({'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white'], 'indent': 1}),
                'rbm_store_name_alt': workbook.add_format({'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'indent': 1}),
                'rbm_conversion_low': workbook.add_format({'font_size': 10, 'font_color': colors_palette['danger_red'], 'bg_color': colors_palette['light_red'], 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['danger_red'], 'num_format': '0.00%', 'bold': True}),
                'rbm_conversion_green': workbook.add_format({'bold': True, 'font_size': 10, 'font_color': colors_palette['success_green'], 'bg_color': colors_palette['light_green'], 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['success_green'], 'num_format': '0.00%'}),
                'rbm_conversion_format': workbook.add_format({'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'num_format': '0.00%'}),
                'rbm_conversion_format_alt': workbook.add_format({'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'num_format': '0.00%'}),
                'rbm_total': workbook.add_format({'bold': True, 'font_size': 12, 'font_color': colors_palette['white'], 'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter', 'border': 2, 'border_color': colors_palette['mint_green']}),
                'rbm_total_label': workbook.add_format({'bold': True, 'font_size': 12, 'font_color': colors_palette['white'], 'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter', 'border': 2, 'border_color': colors_palette['mint_green']}),
                'rbm_summary': workbook.add_format({'bold': True, 'font_size': 10, 'font_color': colors_palette['royal_blue'], 'bg_color': colors_palette['light_royal'], 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['royal_blue']}),
                'rbm_performance': workbook.add_format({'bold': True, 'font_size': 10, 'font_color': colors_palette['white'], 'bg_color': colors_palette['accent_purple'], 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['accent_purple']}),
            }

            ist = pytz.timezone('Asia/Kolkata')
            ist_time = datetime.now(ist)

            # ALL STORES SHEET
            all_data = report_df.sort_values('MTD Value', ascending=False)
            worksheet = workbook.add_worksheet("All Stores")
            headers = ['Store Name', 'FTD Count', 'FTD Value', 'FTD Value Conversion', 'MTD Count', 'MTD Value', 'MTD Value Conversion', 'PREV MONTH SALE', 'DIFF %', 'ASP']
            
            # Column Widths
            column_widths = {}
            for i in range(len(headers)):
                try:
                    if i == 0:
                        max_len = max(all_data[headers[i]].astype(str).map(len).max(), len(headers[i])) + 2
                    else:
                        max_len = max(all_data[headers[i]].map(lambda x: len(str(x))).max() if headers[i] in all_data.columns else 0, len(headers[i])) + 2
                    column_widths[i] = max(max_len, 10)
                except KeyError:
                    column_widths[i] = len(headers[i]) + 2
                worksheet.set_column(i, i, column_widths[i])

            worksheet.merge_range(0, 0, 0, len(headers) - 1, "OSG All Stores Sales Report", formats['title'])
            worksheet.merge_range(1, 0, 1, len(headers) - 1, f"Report Generated: {ist_time.strftime('%d %B %Y %I:%M %p IST')}", formats['subtitle'])

            total_stores = len(all_data)
            active_stores = len(all_data[all_data['FTD Count'] > 0])
            inactive_stores = total_stores - active_stores
            worksheet.merge_range(3, 0, 3, 1, "📊 SUMMARY", formats['header_secondary'])
            worksheet.merge_range(3, 2, 3, len(headers) - 1, f"Total: {total_stores} | Active: {active_stores} | Inactive: {inactive_stores}", formats['data_normal'])

            for col, header in enumerate(headers):
                worksheet.write(5, col, header, formats['header_main'])

            for row_idx, (_, row) in enumerate(all_data.iterrows(), start=6):
                is_alt = (row_idx - 6) % 2 == 1
                worksheet.write(row_idx, 0, row['Store Name'], formats['data_store_name_alt'] if is_alt else formats['data_store_name'])
                worksheet.write(row_idx, 1, int(row['FTD Count']), formats['data_alternate'] if is_alt else formats['data_normal'])
                worksheet.write(row_idx, 2, int(row['FTD Value']), formats['data_alternate'] if is_alt else formats['data_normal'])
                
                ftd_conv = row['FTD Value Conversion']
                fmt = formats['conversion_format_alt'] if is_alt else formats['conversion_format']
                if ftd_conv > 2: fmt = formats['conversion_green']
                elif ftd_conv < 2: fmt = formats['conversion_low']
                worksheet.write(row_idx, 3, ftd_conv/100, fmt)

                worksheet.write(row_idx, 4, int(row['MTD Count']), formats['data_alternate'] if is_alt else formats['data_normal'])
                worksheet.write(row_idx, 5, int(row['MTD Value']), formats['data_alternate'] if is_alt else formats['data_normal'])

                mtd_conv = row['MTD Value Conversion']
                fmt = formats['conversion_format_alt'] if is_alt else formats['conversion_format']
                if mtd_conv > 2: fmt = formats['conversion_green']
                elif mtd_conv < 2: fmt = formats['conversion_low']
                worksheet.write(row_idx, 6, mtd_conv/100, fmt)

                worksheet.write(row_idx, 7, int(row['PREV MONTH SALE']), formats['data_alternate'] if is_alt else formats['data_normal'])
                worksheet.write(row_idx, 8, f"{row['DIFF %']}%", formats['data_alternate'] if is_alt else formats['data_normal'])
                worksheet.write(row_idx, 9, row['ASP'], formats['asp_format_alt'] if is_alt else formats['asp_format'])

            # Total Row
            total_row = len(all_data) + 7
            worksheet.write(total_row, 0, '🎯 TOTAL', formats['total_label'])
            worksheet.write(total_row, 1, all_data['FTD Count'].sum(), formats['total_row'])
            worksheet.write(total_row, 2, all_data['FTD Value'].sum(), formats['total_row'])
            total_ftd_conv = round((all_data['FTD Value'].sum() / all_data['Product_FTD_Amount'].sum()) * 100, 2) if all_data['Product_FTD_Amount'].sum() != 0 else 0
            worksheet.write(total_row, 3, f"{total_ftd_conv}%", formats['total_row'])
            worksheet.write(total_row, 4, all_data['MTD Count'].sum(), formats['total_row'])
            worksheet.write(total_row, 5, all_data['MTD Value'].sum(), formats['total_row'])
            total_mtd_conv = round((all_data['MTD Value'].sum() / all_data['Product_MTD_Amount'].sum()) * 100, 2) if all_data['Product_MTD_Amount'].sum() != 0 else 0
            worksheet.write(total_row, 6, f"{total_mtd_conv}%", formats['total_row'])
            worksheet.write(total_row, 7, all_data['PREV MONTH SALE'].sum(), formats['total_row'])
            total_diff = round(((all_data['MTD Value'].sum() - all_data['PREV MONTH SALE'].sum()) / all_data['PREV MONTH SALE'].sum()) * 100, 2) if all_data['PREV MONTH SALE'].sum() != 0 else 0
            worksheet.write(total_row, 8, f"{total_diff}%", formats['total_row'])
            total_asp = round(all_data['MTD Value'].sum() / all_data['MTD Count'].sum(), 2) if all_data['MTD Count'].sum() != 0 else 0
            worksheet.write(total_row, 9, total_asp, formats['asp_total'])

            if len(all_data) > 0:
                top_performer = all_data.iloc[0]
                insights_row = total_row + 2
                worksheet.merge_range(insights_row, 0, insights_row, len(headers) - 1, f"🏆 Top Performer: {top_performer['Store Name']} (₹{int(top_performer['MTD Value']):,})", formats['data_normal'])

            # RBM SHEETS
            if 'RBM' in report_df.columns:
                rbm_headers = ['Store Name', 'MTD Value Conversion', 'FTD Value Conversion', 'MTD Count', 'FTD Count', 'MTD Value', 'FTD Value', 'PREV MONTH SALE', 'DIFF %', 'ASP']
                for rbm in report_df['RBM'].dropna().unique():
                    if str(rbm) == 'Unknown' or str(rbm) == 'nan': continue
                    
                    rbm_data = report_df[report_df['RBM'] == rbm].sort_values('MTD Value', ascending=False)
                    worksheet_name = str(rbm)[:31]
                    rbm_ws = workbook.add_worksheet(worksheet_name)

                    # Column Widths
                    rbm_column_widths = {}
                    for i in range(len(rbm_headers)):
                        try:
                            if i == 0:
                                max_len = max(rbm_data[rbm_headers[i]].astype(str).map(len).max(), len(rbm_headers[i])) + 2
                            else:
                                max_len = max(rbm_data[rbm_headers[i]].map(lambda x: len(str(x))).max() if rbm_headers[i] in rbm_data.columns else 0, len(rbm_headers[i])) + 2
                            rbm_column_widths[i] = max(max_len, 10)
                        except KeyError:
                            rbm_column_widths[i] = len(rbm_headers[i]) + 2
                        rbm_ws.set_column(i, i, rbm_column_widths[i])

                    rbm_ws.merge_range(0, 0, 0, len(rbm_headers) - 1, f" {rbm} - Sales Performance Report", formats['rbm_title'])
                    rbm_ws.merge_range(1, 0, 1, len(rbm_headers) - 1, f"Report Period: {ist_time.strftime('%B %Y')} | Generated: {ist_time.strftime('%d %B %Y %I:%M %p IST')}", formats['rbm_subtitle'])

                    rbm_total_stores = len(rbm_data)
                    rbm_active_stores = len(rbm_data[rbm_data['FTD Count'] > 0])
                    rbm_inactive_stores = rbm_total_stores - rbm_active_stores
                    rbm_total_amount = rbm_data['MTD Value'].sum()

                    rbm_ws.merge_range(3, 0, 3, 1, "📈 PERFORMANCE OVERVIEW", formats['rbm_summary'])
                    rbm_ws.merge_range(3, 2, 3, len(rbm_headers) - 1, f"Total Stores: {rbm_total_stores} | Active: {rbm_active_stores} | Inactive: {rbm_inactive_stores} | Total Revenue: ₹{rbm_total_amount:,}", formats['rbm_summary'])

                    if len(rbm_data) > 0:
                        best_performer = rbm_data.iloc[0]
                        rbm_ws.merge_range(4, 0, 4, len(rbm_headers) - 1, f"🥇 Best Performer: {best_performer['Store Name']} - ₹{int(best_performer['MTD Value']):,}", formats['rbm_performance'])

                    for col, header in enumerate(rbm_headers):
                        rbm_ws.write(6, col, header, formats['rbm_header'])

                    for row_idx, (_, row) in enumerate(rbm_data.iterrows(), start=7):
                        is_alt = (row_idx - 7) % 2 == 1
                        rbm_ws.write(row_idx, 0, row['Store Name'], formats['rbm_store_name_alt'] if is_alt else formats['rbm_store_name'])
                        
                        mtd_conv = row['MTD Value Conversion']
                        fmt = formats['rbm_conversion_format_alt'] if is_alt else formats['rbm_conversion_format']
                        if mtd_conv > 2: rbm_ws.write(row_idx, 1, mtd_conv/100, formats['rbm_conversion_green'])
                        elif mtd_conv < 2: rbm_ws.write(row_idx, 1, mtd_conv/100, formats['rbm_conversion_low'])
                        else: rbm_ws.write(row_idx, 1, mtd_conv/100, fmt)

                        ftd_conv = row['FTD Value Conversion']
                        if ftd_conv > 2: rbm_ws.write(row_idx, 2, ftd_conv/100, formats['rbm_conversion_green'])
                        elif ftd_conv < 2: rbm_ws.write(row_idx, 2, ftd_conv/100, formats['rbm_conversion_low'])
                        else: rbm_ws.write(row_idx, 2, ftd_conv/100, fmt)

                        data_fmt = formats['rbm_data_alternate'] if is_alt else formats['rbm_data_normal']
                        rbm_ws.write(row_idx, 3, int(row['MTD Count']), data_fmt)
                        rbm_ws.write(row_idx, 4, int(row['FTD Count']), data_fmt)
                        rbm_ws.write(row_idx, 5, int(row['MTD Value']), data_fmt)
                        rbm_ws.write(row_idx, 6, int(row['FTD Value']), data_fmt)
                        rbm_ws.write(row_idx, 7, int(row['PREV MONTH SALE']), data_fmt)
                        rbm_ws.write(row_idx, 8, f"{row['DIFF %']}%", data_fmt)
                        rbm_ws.write(row_idx, 9, row['ASP'], formats['asp_format_alt'] if is_alt else formats['asp_format'])

                    total_row = len(rbm_data) + 8
                    rbm_ws.write(total_row, 0, '🎯 TOTAL', formats['rbm_total_label'])
                    rbm_total_mtd = round((rbm_data['MTD Value'].sum() / rbm_data['Product_MTD_Amount'].sum()) * 100, 2) if rbm_data['Product_MTD_Amount'].sum() != 0 else 0
                    rbm_ws.write(total_row, 1, f"{rbm_total_mtd}%", formats['rbm_total'])
                    rbm_total_ftd = round((rbm_data['FTD Value'].sum() / rbm_data['Product_FTD_Amount'].sum()) * 100, 2) if rbm_data['Product_FTD_Amount'].sum() != 0 else 0
                    rbm_ws.write(total_row, 2, f"{rbm_total_ftd}%", formats['rbm_total'])
                    rbm_ws.write(total_row, 3, rbm_data['MTD Count'].sum(), formats['rbm_total'])
                    rbm_ws.write(total_row, 4, rbm_data['FTD Count'].sum(), formats['rbm_total'])
                    rbm_ws.write(total_row, 5, rbm_data['MTD Value'].sum(), formats['rbm_total'])
                    rbm_ws.write(total_row, 6, rbm_data['FTD Value'].sum(), formats['rbm_total'])
                    rbm_ws.write(total_row, 7, rbm_data['PREV MONTH SALE'].sum(), formats['rbm_total'])
                    
                    total_prev = rbm_data['PREV MONTH SALE'].sum()
                    total_curr = rbm_data['MTD Value'].sum()
                    growth = round(((total_curr - total_prev) / total_prev) * 100, 2) if total_prev != 0 else 0
                    rbm_ws.write(total_row, 8, f"{growth}%", formats['rbm_total'])
                    
                    overall_asp = round(rbm_data['MTD Value'].sum() / rbm_data['MTD Count'].sum(), 2) if rbm_data['MTD Count'].sum() != 0 else 0
                    rbm_ws.write(total_row, 9, overall_asp, formats['asp_total'])

                    insights_row = total_row + 2
                    if growth > 15:
                        rbm_ws.merge_range(insights_row, 0, insights_row, len(rbm_headers) - 1, f"📈 Excellent Growth: {growth}% increase from previous month", formats['rbm_summary'])
                    elif growth < 0:
                        rbm_ws.merge_range(insights_row, 0, insights_row, len(rbm_headers) - 1, f"📉 Needs Attention: {abs(growth)}% decrease from previous month", formats['rbm_summary'])
                    else:
                        rbm_ws.merge_range(insights_row, 0, insights_row, len(rbm_headers) - 1, f"📊 Stable Performance: Less change from previous month", formats['rbm_summary'])

                    insights_row += 1
                    top_3_stores = rbm_data.head(3)
                    if len(top_3_stores) > 0:
                        top_stores_text = " | ".join([f"{store['Store Name']}: ₹{int(store['MTD Value']):,}" for _, store in top_3_stores.iterrows()])
                        rbm_ws.merge_range(insights_row, 0, insights_row, len(rbm_headers) - 1, f"🏆 Top 3 Performers: {top_stores_text}", formats['rbm_summary'])

        excel_output.seek(0)
        return send_file(excel_output, as_attachment=True, download_name=f"OSG_Sales_Report_{datetime.now().strftime('%Y%m%d')}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        import traceback
        return f"ERROR: {traceback.format_exc()}", 500

# ---------------------------------------------------------
# PROCESS: REPORT 2 (DAY VIEW)
# ---------------------------------------------------------

@app.route("/process_report2", methods=["POST"])
def process_report2():
    try:
        report_date = pd.to_datetime(request.form['report_date'])
        time_slot = request.form['time_slot']
        sales_file = request.files['sales_file']

        formatted_date = report_date.strftime("%d-%m-%Y")
        report_title = f"{formatted_date} EW Sale Till {time_slot}"

        future_df = pd.read_excel("Future Store List.xlsx")
        book2_df = pd.read_excel(sales_file)
        book2_df.rename(columns={'Branch': 'Store'}, inplace=True)

        agg = book2_df.groupby('Store', as_index=False).agg({
            'QUANTITY': 'sum',
            'AMOUNT': 'sum'
        })

        all_stores = pd.DataFrame(pd.concat([future_df['Store'], agg['Store']]).unique(), columns=['Store'])
        merged = all_stores.merge(agg, on='Store', how='left')
        merged['QUANTITY'] = merged['QUANTITY'].fillna(0).astype(int)
        merged['AMOUNT'] = merged['AMOUNT'].fillna(0).astype(int)

        merged = merged.sort_values(by='AMOUNT', ascending=False).reset_index(drop=True)

        total = pd.DataFrame([{
            'Store': 'TOTAL',
            'QUANTITY': merged['QUANTITY'].sum(),
            'AMOUNT': merged['AMOUNT'].sum()
        }])

        final_df = pd.concat([merged, total], ignore_index=True)
        final_df.rename(columns={'Store': 'Branch'}, inplace=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Store Report"

        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = report_title
        title_cell.font = Font(bold=True, size=11, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill("solid", fgColor="4F81BD")

        header_fill = PatternFill("solid", fgColor="4F81BD")
        data_fill = PatternFill("solid", fgColor="DCE6F1")
        red_fill = PatternFill("solid", fgColor="F4CCCC")
        total_fill = PatternFill("solid", fgColor="10B981")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, color="FFFFFF")
        bold_font = Font(bold=True, color="FFFFFF")

        for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 2:
                    cell.fill = header_fill
                    cell.font = header_font
                elif final_df.loc[r_idx - 3, 'Branch'] == 'TOTAL':
                    cell.fill = total_fill
                    cell.font = bold_font
                elif final_df.loc[r_idx - 3, 'AMOUNT'] <= 0:
                    cell.fill = red_fill
                else:
                    cell.fill = data_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name=f"Store_Summary_{formatted_date}_{time_slot}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print("ERROR IN REPORT 2:")
        print(error_details)
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Error - Day View Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; padding: 2rem; background: #f8f9fa; }}
                .error-container {{ background: white; padding: 2rem; border-radius: 8px; max-width: 1000px; margin: 0 auto; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                h1 {{ color: #dc3545; }}
                pre {{ background: #f8f9fa; padding: 1rem; border-radius: 4px; overflow-x: auto; border-left: 4px solid #dc3545; }}
                .error-msg {{ color: #721c24; background: #f8d7da; padding: 1rem; border-radius: 4px; margin: 1rem 0; }}
            </style>
        </head>
        <body>
            <div class="error-container">
                <h1>⚠️ Error Processing Day View Report</h1>
                <div class="error-msg">
                    <strong>Error:</strong> {str(e)}
                </div>
                <h3>Full Stack Trace:</h3>
                <pre>{error_details}</pre>
                <p><a href="/report2">← Go Back</a></p>
            </div>
        </body>
        </html>
        """, 500

# ---------------------------------------------------------
# PROCESS: WARRANTY CLAIM MANAGEMENT
# ---------------------------------------------------------

@app.route("/warranty")
def warranty_page():
    return render_template("warranty.html")

@app.route("/warranty/lookup", methods=["POST"])
def warranty_lookup():
    try:
        data = request.get_json()
        mobile = data.get("mobile", "").strip()
        
        if not mobile or len(mobile) != 10:
            return {"found": False, "message": "Invalid mobile number"}, 400

        # Use the stateless loader from claim_processor
        df = claim_processor.load_excel_data()
        customer_data = claim_processor.get_customer_records(df, mobile)

        if customer_data.empty:
            return {"found": False, "message": "No records found"}, 200

        # Extract customer name safely - prioritize "name" column
        customer_col = claim_processor.resolve_column(df, ["name", "customer name", "customer"])
        customer_name = str(customer_data.iloc[0].get(customer_col, "Unknown"))

        # Build product list
        products = []
        invoice_col = claim_processor.resolve_column(df, ["invoice no", "invoice", "invoice_no"])
        model_col = claim_processor.resolve_column(df, ["model"])
        serial_col = claim_processor.resolve_column(df, ["serial no", "serialno", "serial_no"])
        osid_col = claim_processor.resolve_column(df, ["osid"])

        for _, row in customer_data.iterrows():
            products.append({
                "invoice": str(row.get(invoice_col, "")),
                "model": str(row.get(model_col, "")),
                "serial": str(row.get(serial_col, "")),
                "osid": str(row.get(osid_col, ""))
            })

        return {
            "found": True,
            "name": customer_name,
            "products": products
        }

    except Exception as e:
        print(f"Error in warranty lookup: {e}", file=sys.stderr)
        return {"found": False, "message": str(e)}, 500

@app.route("/warranty/submit", methods=["POST"])
def warranty_submit():
    try:
        mobile = request.form.get("mobile_no")
        address = request.form.get("address")
        issue_desc = request.form.get("issue_description")
        products_json = request.form.get("products_json")
        
        # Handle file upload
        uploaded_file = request.files.get("document")
        file_path = None
        if uploaded_file and uploaded_file.filename:
            # Save temporarily
            import os
            temp_dir = "temp_uploads"
            os.makedirs(temp_dir, exist_ok=True)
            file_path = os.path.join(temp_dir, uploaded_file.filename)
            uploaded_file.save(file_path)

        if not products_json:
            return {"success": False, "message": "No products selected"}, 400

        selected_products = json.loads(products_json)

        # ---------------------------------------------------------
        # OPTIMIZATION: Background Processing & Instant Cache Update
        # ---------------------------------------------------------
        
        # 1. Update Tracking Cache Immediately (Optimistic Update)
        global _TRACKING_CACHE, _TRACKING_CACHE_TIME
        
        # Get customer name for the cache
        try:
            df = claim_processor.load_excel_data()
            customer_records = claim_processor.get_customer_records(df, mobile)
            if not customer_records.empty:
                name_col = claim_processor.resolve_column(df, ["name", "customer name", "customer"])
                customer_name = str(customer_records.iloc[0].get(name_col, "Customer"))
            else:
                customer_name = "Customer"
        except:
            customer_name = "Customer"

        new_claim = {
            "customer_name": customer_name,
            "mobile_no": mobile,
            "address": address,
            "products": "; ".join([p.get("invoice", "") for p in selected_products]),
            "issue_description": issue_desc,
            "status": "Pending (Processing)", # Show distinct status
            "submitted_date": datetime.now().isoformat()
        }
        
        # Initialize cache if empty
        if _TRACKING_CACHE is None:
            _TRACKING_CACHE = []
            
        # Add new claim to the TOP of the list
        _TRACKING_CACHE.insert(0, new_claim)
        # Extend cache TTL so this new data sticks around
        _TRACKING_CACHE_TIME = time.time() 
        
        # 2. Run Heavy Processing in Background Thread
        def process_in_background(mob, addr, prods, issue, fpath):
            try:
                print(f"Starting background processing for {mob}", file=sys.stderr)
                claim_processor.process_claim(
                    mobile=mob,
                    address=addr,
                    selected_products=prods,
                    global_issue=issue,
                    global_file_path=fpath
                )
                print(f"Background processing completed for {mob}", file=sys.stderr)
                
                # Update status in cache to "Pending" (remove Processing tag)
                for c in _TRACKING_CACHE:
                    if c.get("mobile_no") == mob and c.get("issue_description") == issue:
                        c["status"] = "Pending"
                        break
                        
                # Cleanup temp file
                if fpath and os.path.exists(fpath):
                    os.remove(fpath)
                        
            except Exception as e:
                print(f"Background processing failed: {e}", file=sys.stderr)
                # Cleanup temp file on error too
                if fpath and os.path.exists(fpath):
                    os.remove(fpath)

        import threading
        threading.Thread(
            target=process_in_background,
            args=(mobile, address, selected_products, issue_desc, file_path),
            daemon=True
        ).start()

        # 3. Return Success Immediately
        return {"success": True, "message": "Claim submitted successfully! Processing in background."}

    except Exception as e:
        print(f"Error in warranty submit: {e}", file=sys.stderr)
        return {"success": False, "message": str(e)}, 500

# Cache for tracking data
_TRACKING_CACHE = None
_TRACKING_CACHE_TIME = 0
_TRACKING_CACHE_TTL = 60  # seconds

@app.route("/warranty/track-data")
def warranty_track_data():
    global _TRACKING_CACHE, _TRACKING_CACHE_TIME
    
    try:
        # Check cache
        current_time = time.time()
        if _TRACKING_CACHE is not None and (current_time - _TRACKING_CACHE_TIME) < _TRACKING_CACHE_TTL:
            print("Serving tracking data from cache", file=sys.stderr)
            all_claims = _TRACKING_CACHE
        else:
            # Proxy the request to Google Script to avoid CORS issues on client
            import requests
            print(f"Fetching claims from: {claim_processor.WEB_APP_URL}", file=sys.stderr)
            response = requests.get(claim_processor.WEB_APP_URL, timeout=10)
            print(f"Response status: {response.status_code}", file=sys.stderr)
            
            if response.status_code == 200:
                try:
                    all_claims = response.json()
                    print(f"Retrieved {len(all_claims) if isinstance(all_claims, list) else 0} claims", file=sys.stderr)
                    
                    # Normalize column names - Google Sheets uses "Mobile No" but we need "mobile_no"
                    normalized_claims = []
                    for claim in all_claims:
                        normalized = {}
                        for key, value in claim.items():
                            # Normalize keys to lowercase with underscores
                            normalized_key = key.lower().replace(" ", "_")
                            normalized[normalized_key] = value
                            # Also keep original key for backward compatibility
                            normalized[key] = value
                        normalized_claims.append(normalized)
                    
                    # Update cache
                    all_claims = normalized_claims
                    _TRACKING_CACHE = all_claims
                    _TRACKING_CACHE_TIME = current_time
                    print(f"Normalized and cached {len(normalized_claims)} claims", file=sys.stderr)
                    
                except Exception as json_err:
                    print(f"JSON parse error: {json_err}", file=sys.stderr)
                    print(f"Response text: {response.text[:200]}", file=sys.stderr)
                    return json.dumps([]), 200, {'Content-Type': 'application/json'}
            else:
                print(f"Non-200 status code: {response.status_code}, Response: {response.text[:200]}", file=sys.stderr)
                return json.dumps([]), 200, {'Content-Type': 'application/json'}

        # Filter if mobile provided
        mobile = request.args.get("mobile", "").strip()
        if mobile:
            print(f"Filtering by mobile: {mobile}", file=sys.stderr)
            # Check both normalized and original column names
            filtered = [
                c for c in all_claims 
                if str(c.get("mobile_no", "")).strip() == mobile 
                or str(c.get("Mobile No", "")).strip() == mobile
            ]
            print(f"Found {len(filtered)} matching claims", file=sys.stderr)
            return json.dumps(filtered), 200, {'Content-Type': 'application/json'}
        
        return json.dumps(all_claims), 200, {'Content-Type': 'application/json'}

    except Exception as e:
        print(f"Error fetching track data: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        return json.dumps([]), 200, {'Content-Type': 'application/json'}

if __name__ == "__main__":
    # Pre-load data in a background thread to avoid blocking startup
    import threading
    def preload_data():
        try:
            print("Pre-loading Excel data...", file=sys.stderr)
            claim_processor.load_excel_data()
            print("Excel data pre-loaded successfully.", file=sys.stderr)
        except Exception as e:
            print(f"Failed to pre-load data: {e}", file=sys.stderr)
            
    threading.Thread(target=preload_data, daemon=True).start()
    
    app.run()
